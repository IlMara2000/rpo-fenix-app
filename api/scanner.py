from http.server import BaseHTTPRequestHandler
import cgi
import io
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={"REQUEST_METHOD": "POST"},
            )

            file_excel = form["excel"].file.read()
            file_txt = form["txt"].file.read().decode("utf-8")
            blacklist = set()

            for line in file_txt.splitlines():
                match = re.search(r"^(\d{9,10})", line.strip())
                if match:
                    blacklist.add(match.group(1))

            if not blacklist:
                raise ValueError("Nessun numero di telefono valido trovato nel file TXT.")

            wb = load_workbook(io.BytesIO(file_excel))
            ws = wb.active
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            black_font = Font(color="000000")
            matches_count = 0

            for row in ws.iter_rows(min_row=1):
                found_in_blacklist = False
                for cell in row:
                    if not cell.value:
                        continue

                    clean_value = "".join(filter(str.isdigit, str(cell.value)))
                    if clean_value.startswith("39") and len(clean_value) > 10:
                        clean_value = clean_value[2:]

                    if clean_value in blacklist:
                        found_in_blacklist = True
                        break

                if found_in_blacklist:
                    matches_count += 1
                    for cell in row:
                        cell.fill = black_fill
                        cell.font = black_font

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            self.send_response(200)
            self.send_header(
                "Content-type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("X-Matches", str(matches_count))
            self.end_headers()
            self.wfile.write(output.read())
        except Exception as error:
            self.send_response(500)
            self.end_headers()
            self.wfile.write(f"Errore: {str(error)}".encode())
